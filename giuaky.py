from openpyxl import load_workbook
import os
from os import listdir

# import module
import traceback
path = 'doc'


conmment = {
    10: "Biết vận dụng tốt từ vựng và các mẫu câu đã học.",

    9: "Nắm vững cấu trúc câu, nhớ và hiểu từ vựng.",

    8:  "Các kỹ năng nghe, nói, đọc, viết tiến bộ.",

    7: "Hoàn thành nội dung các bài học.",

    6: "Hoàn thành nội dung các bài học.",

    5: "Nắm được nội dung bài học."

}


symbol = {
    "T": "Hoàn thành tốt nội dung bài học. Nghe, nói , đọc, viết tốt",
    "H": "Nắm được cấu trúc câu, hiểu từ các từ vựng",
    "C":  "Sử dụng mẫu câu và từ vựng còn hạn chế"
}


for file in listdir(path):
    wb = load_workbook(os.path.join(path, file))
    ws = wb.active
    for row in range(2, ws.max_row+1):
        try:
            content_ref = ws.cell(row, 8)
            symbol_ref = ws.cell(row, 5)

            try:
                content_ref.value = symbol[symbol_ref.value]

            except Exception as e:
                traceback.print_exc()
                content_ref.value = ''

        except:
            print("Loi diem", os.path.join(path, file))

    wb.save(os.path.join(path, file))
