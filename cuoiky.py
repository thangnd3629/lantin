from openpyxl import load_workbook
import os
from os import listdir
path = 'doc'


conmment = {
    10: "Biết vận dụng tốt từ vựng và các mẫu câu đã học.",

    9: "Nắm vững cấu trúc câu, nhớ và hiểu từ vựng.",

    8:  "Các kỹ năng nghe, nói, đọc, viết tiến bộ.",

    7: "Hoàn thành nội dung các bài học.",

    6: "Hoàn thành nội dung các bài học.",

    5: "Nắm được nội dung bài học."

}


symbol = {
    10: "T",
    9: "T",
    8:  "H",
    6: "H",
    7: "H",
    5: "H"
}


for file in listdir(path):
    wb = load_workbook(os.path.join(path, file))
    ws = wb.active
    for row in range(2, ws.max_row+1):

        assesment = ws.cell(row, 6).value
        try:
            ws.cell(row, 7).value = ""
            assesment = int(assesment)
            content_ref = ws.cell(row, 8)
            symbol_ref = ws.cell(row, 5)

            try:
                if assesment < 5:
                    content_ref.value = "Chưa hoàn thành nội dung học kỳ I"
                    symbol_ref.value = "C"
                else:
                    content_ref.value = conmment[assesment]
                    symbol_ref.value = symbol[assesment]

            except:

                content_ref.value = ''

        except:
            print("Loi diem", os.path.join(path, file), " line ", row)

    wb.save(os.path.join(path, file))
